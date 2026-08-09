# -*- coding: utf-8 -*-
""" 【A股强势股扫描器 v81_MinCredible Final】（Gemini打磨 + 单位/放量比修复）
基于 v7.9 / v81：保留自选 Excel + 七维评分 + 全部原有技术指标算法。
v81 最终打磨版：
1) 最新价/获利%：spot无效回退K线收盘（消灭-100%计算异常）
2) 缓存 TTL（默认2小时），避免脏缓存与复权因子混用
3) 历史基准日：周末/非交易日自动回退至上周五
4) RS merge 对齐：对齐交易日计算相对沪深300强度
5) 池内分位：总分分位%、RS分位%（增强小样本区分度）
6) 大盘MA60环境写入建议；BIAS高位防追高
7) 基本面/资金流默认关闭（ Actions 运行极速且稳定）
8) 周末：强制以最近交易日K线收盘回填价格/涨跌/振幅/成交量，避免数据断层
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from pathlib import Path
import sys
import re
import time
import warnings
import traceback
import random
import threading
import pickle
import shutil
from concurrent.futures import ThreadPoolExecutor, as_completed
from threading import Lock

warnings.filterwarnings('ignore')

try:
    from tqdm import tqdm
    USE_TQDM = True
except ImportError:
    USE_TQDM = False
    print("⚠️ 未安装 tqdm，将不显示进度条。建议运行: pip install tqdm")

import akshare as ak
import baostock as bs

# ================== 配置区域 ==================
POSSIBLE_INPUTS = [
    Path("输入股票代码及名称清单v1.xlsx"),      # ← 默认配置文件名
]

OUTPUT_DIR = Path("results")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

CACHE_DIR = Path("stock_cache_v81")
CACHE_DIR.mkdir(parents=True, exist_ok=True)
CACHE_TTL_HOURS = 2  # 缓存有效期（小时）
CACHE_LOCK = threading.Lock()

MAX_HIST_DAYS = 150
MAX_WORKERS = 1  # 自选约10只：单线程最稳，防数据源限流
MIN_HIST_DAYS = 30
BAOSTOCK_RETRY = 5

MARKET_CONFIG = {
    '主板': {'limit_pct': 0.095, 'breakout_pct': 6.0, 'atr_mult_high': 3.0, 'atr_mult_mid': 2.5, 'atr_mult_low': 2.0, 'max_drop': 0.20},
    '科创/创业板': {'limit_pct': 0.195, 'breakout_pct': 8.0, 'atr_mult_high': 3.5, 'atr_mult_mid': 3.0, 'atr_mult_low': 2.5, 'max_drop': 0.30},
    '北交所': {'limit_pct': 0.295, 'breakout_pct': 10.0, 'atr_mult_high': 4.0, 'atr_mult_mid': 3.5, 'atr_mult_low': 3.0, 'max_drop': 0.40}
}
EPSILON = 1e-9
BIAS_HIGH_THRESHOLD = 5.0
VOL_BOOST_THRESHOLD = 1.5
DEFAULT_ATR_PERIOD = 14

BS_LOCK = Lock()

# ================== 多数据源回退 ==================
BS_LOGGED_IN = False
BENCHMARK_NAME = "未知"  # 相对强度基准指数名称

def bs_login_once():
    global BS_LOGGED_IN
    with BS_LOCK:
        if not BS_LOGGED_IN:
            try:
                bs.login()
                BS_LOGGED_IN = True
                print("✅ Baostock 登录成功")
                return True
            except Exception as e:
                print(f"❌ Baostock 登录失败: {e}")
                return False
        return True

def get_baostock_symbol(code: str) -> str:
    code = re.sub(r'\D', '', str(code).strip())
    if len(code) != 6:
        return ""
    if code.startswith(('6', '5', '9')):
        return f"sh.{code}"
    return f"sz.{code}"

def fetch_with_baostock(code: str, start_date: str, end_date: str):
    if not bs_login_once():
        return None
    symbol = get_baostock_symbol(code)
    if not symbol:
        return None

    print(f"🔍 Baostock 查询 {symbol} | 开始: {start_date} | 结束: {end_date}")

    for attempt in range(BAOSTOCK_RETRY):
        try:
            with BS_LOCK:
                rs = bs.query_history_k_data_plus(
                    symbol,
                    "date,open,high,low,close,volume,amount,turn,peTTM,pbMRQ",
                    start_date=start_date,
                    end_date=end_date,
                    frequency="d",
                    adjustflag="2"  # 固定前复权
                )
                if rs.error_code != '0':
                    raise RuntimeError(f"baostock error_code={rs.error_code}")
                df = rs.get_data()
            if df is None or df.empty or len(df) < MIN_HIST_DAYS:
                raise RuntimeError("empty or too short")
            print(f"   ✅ 第{attempt+1}次成功 获取 {len(df)} 条数据 (qfq)")
            df = df.astype({'open':'float','high':'float','low':'float','close':'float',
                            'volume':'float','amount':'float','turn':'float'})
            df.rename(columns={'turn': 'turnover_rate'}, inplace=True)
            df['date'] = pd.to_datetime(df['date'])
            return df
        except Exception as e:
            print(f"   ❌ 第{attempt+1}次失败: {str(e)[:120]}")
            time.sleep(1.5 + random.uniform(0, 1.5))
    print(f"❌ {code} Baostock 所有尝试均失败")
    return None

def fetch_hist_with_fallback(code: str, end_date_str: str):
    symbol = re.sub(r'\D', '', str(code).strip().upper())
    start_date_ak = (datetime.now() - timedelta(days=MAX_HIST_DAYS)).strftime("%Y%m%d")
    end_str_ak = end_date_str.replace('-', '')
    start_date_bs = (datetime.now() - timedelta(days=MAX_HIST_DAYS)).strftime("%Y-%m-%d")
    end_date_bs = end_date_str

    for attempt in range(2):
        try:
            df = ak.stock_zh_a_hist(symbol=symbol, period="daily",
                                    start_date=start_date_ak, end_date=end_str_ak, adjust="qfq")
            if not df.empty and len(df) >= MIN_HIST_DAYS:
                df = df.rename(columns={'日期':'date','开盘':'open','最高':'high',
                                        '最低':'low','收盘':'close','成交量':'volume'})
                return df
            time.sleep(0.5)
        except Exception:
            time.sleep(0.8 + random.uniform(0, 0.3))

    print(f"  └─ {code} akshare 失败，切换 Baostock...")
    df = fetch_with_baostock(code, start_date_bs, end_date_bs)
    if df is not None and len(df) >= MIN_HIST_DAYS:
        return df
    return None

# ================== 基础工具 ==================
def get_last_trade_day():
    """最近已收盘的交易日（周末/周一开盘前：自动回退到周五）。"""
    d = datetime.now()
    if d.weekday() >= 5:
        d -= timedelta(days=d.weekday() - 4)
        return d.strftime("%Y-%m-%d")
    for _ in range(10):
        if d.weekday() < 5:
            return d.strftime("%Y-%m-%d")
        d -= timedelta(days=1)
    return datetime.now().strftime("%Y-%m-%d")

def is_weekend_session() -> bool:
    """当前是否周末（无盘），应用最近交易日收盘数据。"""
    return datetime.now().weekday() >= 5

def session_data_note(end_date_str: str) -> str:
    if is_weekend_session():
        return f"周末无盘，分析基准日={end_date_str}（上周五收盘），实时字段可能为0，价格类指标沿用该日收盘"
    return f"分析基准日={end_date_str}"

def is_beijing_stock(code):
    code = re.sub(r'\D', '', str(code).strip())
    if len(code) != 6:
        return False
    return code[:2] in ['83', '87', '88', '89'] or code[:3] == '920'

def detect_market_type(code):
    code = re.sub(r'\D', '', str(code).strip())
    if not code or len(code) != 6:
        return '主板'
    if is_beijing_stock(code):
        return '北交所'
    if code.startswith('3') or code.startswith('688') or code.startswith('689'):
        return '科创/创业板'
    return '主板'

def get_akshare_symbol(code):
    return re.sub(r'\D', '', str(code).strip().upper())

def clean_numeric(df):
    df.columns = [str(c).lower().strip() for c in df.columns]
    price_map = {'开盘':'open','最高':'high','最低':'low','收盘':'close',
                 '开盘价':'open','最高价':'high','最低价':'low','收盘价':'close','日期':'date'}
    for cn, en in price_map.items():
        if cn in df.columns and en not in df.columns:
            df[en] = df[cn]
    for col in ['open', 'high', 'low', 'close']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')
    vol_cols = ['成交量', 'volume', 'vol', 'vol_amt']
    for col in vol_cols:
        if col in df.columns:
            df['volume'] = pd.to_numeric(df[col], errors='coerce').fillna(0)
            break
    if 'volume' not in df.columns:
        df['volume'] = 0.0
    turn_cols = ['换手率', 'turnover_rate', 'turnover', '周转率', 'hsl']
    for col in turn_cols:
        if col in df.columns:
            df['turnover_rate'] = pd.to_numeric(df[col], errors='coerce').fillna(0)
            break
    if 'turnover_rate' not in df.columns:
        df['turnover_rate'] = 0.0
    if 'date' not in df.columns:
        df['date'] = pd.to_datetime(df.index)
    else:
        df['date'] = pd.to_datetime(df['date'])
    df = df.dropna(subset=['close'])
    df = df[df['close'] > EPSILON]
    return df

def calc_stock_mas(hist):
    out = {'MA5': 0.0, 'MA7': 0.0, 'MA10': 0.0, 'MA20': 0.0, 'MA60': 0.0, 'MA120': 0.0}
    if hist is None or len(hist) < 5:
        return out
    c = hist['close']
    for n, key in [(5, 'MA5'), (7, 'MA7'), (10, 'MA10'), (20, 'MA20'), (60, 'MA60'), (120, 'MA120')]:
        if len(c) >= n:
            out[key] = round(float(c.tail(n).mean()), 3)
    return out


def classify_adx_state(adx):
    """ADX 趋势强度 → 市场状态（经典阈值）"""
    try:
        v = float(adx)
    except Exception:
        return '未知'
    if v < 20:
        return '无趋势'
    if v < 25:
        return '趋势形成中'
    if v < 40:
        return '强趋势'
    if v <= 50:
        return '极强趋势'
    return '超强趋势'


def classify_stop_distance_state(dist_pct):
    """止损距离% → 市场状态（风控舒适度）"""
    try:
        v = float(dist_pct)
    except Exception:
        return '未知'
    if v < 3:
        return '极低波动盘整'
    if v < 5:
        return '低波动平稳期'
    if v <= 10:
        return '健康趋势行情'
    if v <= 15:
        return '波动加大/震荡期'
    return '剧烈波动/深回撤中'


def calc_bias_bundle(hist, close_price):
    """
    多周期乖离率 + 市场状态
    BIAS(N)=(收盘-MA(N))/MA(N)*100
    阈值参考：BIAS5/10/20/60 对照表
    """
    out = {
        'BIAS5%': 0.0, 'BIAS5状态': '未知',
        'BIAS10%': 0.0, 'BIAS10状态': '未知',
        'BIAS20%': 0.0, 'BIAS20状态': '未知',
        'BIAS60%': 0.0, 'BIAS60状态': '未知',
    }
    if hist is None or close_price is None or close_price <= EPSILON:
        return out
    c = hist['close']

    def bias_n(n):
        if len(c) < n:
            return None
        ma = float(c.tail(n).mean())
        if ma <= EPSILON:
            return None
        return (float(close_price) / ma - 1.0) * 100.0

    def state5(b):
        if b is None:
            return '未知'
        if b < -5:
            return '极度超卖'
        if b < -3:
            return '超卖偏弱'
        if b <= 3:
            return '中性区间'
        if b <= 5:
            return '超买偏强'
        return '极度超买'

    def state10(b):
        if b is None:
            return '未知'
        if b < -8:
            return '极度超卖'
        if b < -5:
            return '超卖偏弱'
        if b <= 5:
            return '中性区间'
        if b <= 8:
            return '超买偏强'
        return '极度超买'

    def state20(b):
        if b is None:
            return '未知'
        if b < -12:
            return '极度超卖'
        if b < -8:
            return '超卖偏弱'
        if b <= 5:
            return '中性区间'
        if b <= 12:
            return '超买偏强'
        return '极度超买'

    def state60(b):
        if b is None:
            return '未知'
        if b < -20:
            return '极度超卖'
        if b < -15:
            return '超卖偏弱'
        if b <= 15:
            return '中性区间'
        if b <= 20:
            return '超买偏强'
        return '极度超买'

    b5 = bias_n(5)
    b10 = bias_n(10)
    b20 = bias_n(20)
    b60 = bias_n(60)
    if b5 is not None:
        out['BIAS5%'] = round(b5, 2)
        out['BIAS5状态'] = state5(b5)
    if b10 is not None:
        out['BIAS10%'] = round(b10, 2)
        out['BIAS10状态'] = state10(b10)
    if b20 is not None:
        out['BIAS20%'] = round(b20, 2)
        out['BIAS20状态'] = state20(b20)
    if b60 is not None:
        out['BIAS60%'] = round(b60, 2)
        out['BIAS60状态'] = state60(b60)
    return out

def _normalize_index_df(df):
    """统一指数列名并清洗"""
    if df is None or getattr(df, 'empty', True):
        return None
    rename = {
        '日期': 'date', 'date': 'date',
        '开盘': 'open', 'open': 'open',
        '最高': 'high', 'high': 'high',
        '最低': 'low', 'low': 'low',
        '收盘': 'close', 'close': 'close',
        '成交量': 'volume', 'volume': 'volume',
    }
    cols = {}
    for c in df.columns:
        cs = str(c).strip()
        if cs in rename:
            cols[c] = rename[cs]
    df = df.rename(columns=cols)
    df = clean_numeric(df)
    if df is None or df.empty or 'close' not in df.columns:
        return None
    if len(df) > 400:
        df = df.tail(400).reset_index(drop=True)
    return df


def fetch_index_hist(symbol, end_date_str, name=''):
    """
    拉取指数日K，多源回退：
    1) 东财 index_zh_a_hist（必须带 start/end）
    2) 新浪 stock_zh_index_daily（sh000001 / sz399001）
    """
    label = name or symbol
    start = (datetime.now() - timedelta(days=500)).strftime("%Y%m%d")
    end = (end_date_str or datetime.now().strftime("%Y-%m-%d")).replace('-', '')
    pure = re.sub(r'\D', '', str(symbol)).zfill(6)

    # 1) 东财
    try:
        df = ak.index_zh_a_hist(symbol=pure, period="daily", start_date=start, end_date=end)
        df = _normalize_index_df(df)
        if df is not None and len(df) >= 30:
            print(f"  ✅ 指数 {label}({pure}) 东财成功 {len(df)} 条")
            return df
    except Exception as e:
        print(f"  ⚠️ 指数 {label} 东财失败: {type(e).__name__}: {str(e)[:60]}")

    # 2) 新浪：上证 sh000001 / 深证 sz399001 / 沪深300 sh000300
    sina_map = {
        '000001': 'sh000001',
        '399001': 'sz399001',
        '000300': 'sh000300',
        '399006': 'sz399006',
    }
    sina_sym = sina_map.get(pure)
    if not sina_sym:
        if pure.startswith('399'):
            sina_sym = f'sz{pure}'
        else:
            sina_sym = f'sh{pure}'
    try:
        df = ak.stock_zh_index_daily(symbol=sina_sym)
        df = _normalize_index_df(df)
        if df is not None and len(df) >= 30:
            print(f"  ✅ 指数 {label}({sina_sym}) 新浪成功 {len(df)} 条")
            return df
    except Exception as e:
        print(f"  ⚠️ 指数 {label} 新浪失败: {type(e).__name__}: {str(e)[:60]}")

    print(f"  ❌ 指数 {label} 全部来源失败")
    return None


def build_market_profile_row(hist, name, code):
    """大盘画像单行：报价摘要 + 均线 + 支撑压力 + 形态"""
    row = {
        '指数名称': name,
        '指数代码': code,
        '最新价': 0.0,
        '昨收': 0.0,
        '今开': 0.0,
        '最高': 0.0,
        '最低': 0.0,
        '涨跌幅%': 0.0,
        '振幅%': 0.0,
        '5日均线': 0.0,
        '10日均线': 0.0,
        '30日均线': 0.0,
        '60日均线': 0.0,
        '120日均线': 0.0,
        '短线支撑位': 0.0,
        '短线压力位': 0.0,
        '超短线支撑位': 0.0,
        '超短线压力位': 0.0,
        'MACD金叉': '无',
        '技术形态': '',
        'K线形态': '',
        '大阳次数': 0,
        '相对MA60': '',
        '趋势结论': '',
    }
    if hist is None or len(hist) < 5:
        row['趋势结论'] = '数据不足'
        return row

    last = hist.iloc[-1]
    close = float(last['close'])
    open_p = float(last['open']) if 'open' in hist.columns else close
    high = float(last['high']) if 'high' in hist.columns else close
    low = float(last['low']) if 'low' in hist.columns else close
    prev = float(hist['close'].iloc[-2]) if len(hist) >= 2 else close
    row['最新价'] = round(close, 2)
    row['昨收'] = round(prev, 2)
    row['今开'] = round(open_p, 2)
    row['最高'] = round(high, 2)
    row['最低'] = round(low, 2)
    if prev > EPSILON:
        row['涨跌幅%'] = round((close / prev - 1) * 100, 2)
        row['振幅%'] = round((high - low) / prev * 100, 2)

    c = hist['close']
    for n, key in [(5, '5日均线'), (10, '10日均线'), (30, '30日均线'), (60, '60日均线'), (120, '120日均线')]:
        if len(c) >= n:
            row[key] = round(float(c.tail(n).mean()), 3)

    try:
        sup, res, ssup, sres = calculate_support_resistance(hist)
        row['短线支撑位'] = sup
        row['短线压力位'] = res
        row['超短线支撑位'] = ssup
        row['超短线压力位'] = sres
    except Exception:
        pass
    try:
        row['MACD金叉'] = check_macd_golden_cross(hist)
    except Exception:
        pass
    try:
        hist_ind = precompute_indicators(hist)
        row['技术形态'] = detect_technical_patterns(hist_ind)
    except Exception:
        pass
    try:
        row['K线形态'] = detect_kline_patterns(hist)
    except Exception:
        pass
    try:
        df = hist.tail(61).copy()
        if len(df) >= 2:
            prev_c = df['close'].shift(1)
            pct = (df['close'] / prev_c - 1) * 100
            rng = (df['high'] - df['low']).replace(0, np.nan)
            pos = (df['close'] - df['low']) / rng
            big = ((pct >= 2.0) & (pos >= 0.7)).fillna(False).sum()
            row['大阳次数'] = int(big)
    except Exception:
        pass

    ma60 = row['60日均线']
    if ma60 and ma60 > EPSILON:
        if close > ma60 * 1.002:
            row['相对MA60'] = '站上MA60'
        elif close < ma60 * 0.998:
            row['相对MA60'] = '跌破MA60'
        else:
            row['相对MA60'] = '贴近MA60'
    # 简单趋势结论
    parts = []
    if row['相对MA60']:
        parts.append(row['相对MA60'])
    if row['技术形态'] and row['技术形态'] != '无明显形态':
        parts.append(row['技术形态'])
    if row['MACD金叉'] and row['MACD金叉'] != '否':
        parts.append(str(row['MACD金叉']))
    if row['涨跌幅%'] > 1:
        parts.append('日内偏强')
    elif row['涨跌幅%'] < -1:
        parts.append('日内偏弱')
    row['趋势结论'] = ' | '.join(parts) if parts else '中性观望'
    return row


def build_index_snapshot(hist, prefix):
    out = {}
    ma_map = [(5, '5日均线'), (10, '10日均线'), (30, '30日均线'), (60, '60日均线'), (120, '120日均线')]
    for n, label in ma_map:
        out[f'{prefix}{label}'] = 0.0
    out[f'{prefix}短线支撑位'] = 0.0
    out[f'{prefix}短线压力位'] = 0.0
    out[f'{prefix}超短线支撑位'] = 0.0
    out[f'{prefix}超短线压力位'] = 0.0
    out[f'{prefix}MACD金叉'] = '无'
    out[f'{prefix}技术形态'] = ''
    out[f'{prefix}K线形态'] = ''
    out[f'{prefix}大阳次数'] = 0

    if hist is None or len(hist) < 20:
        return out

    c = hist['close']
    for n, label in ma_map:
        if len(c) >= n:
            out[f'{prefix}{label}'] = round(float(c.tail(n).mean()), 3)

    try:
        sup, res, ssup, sres = calculate_support_resistance(hist)
        out[f'{prefix}短线支撑位'] = sup
        out[f'{prefix}短线压力位'] = res
        out[f'{prefix}超短线支撑位'] = ssup
        out[f'{prefix}超短线压力位'] = sres
    except Exception:
        pass
    try:
        out[f'{prefix}MACD金叉'] = check_macd_golden_cross(hist)
    except Exception:
        out[f'{prefix}MACD金叉'] = '无'
    try:
        # 技术形态依赖 ma5/ma10/ma20，需先预计算
        hist_ind = precompute_indicators(hist)
        out[f'{prefix}技术形态'] = detect_technical_patterns(hist_ind)
    except Exception:
        out[f'{prefix}技术形态'] = ''
    try:
        out[f'{prefix}K线形态'] = detect_kline_patterns(hist)
    except Exception:
        out[f'{prefix}K线形态'] = ''
    try:
        df = hist.tail(61).copy()
        if len(df) >= 2:
            prev = df['close'].shift(1)
            pct = (df['close'] / prev - 1) * 100
            rng = (df['high'] - df['low']).replace(0, np.nan)
            pos = (df['close'] - df['low']) / rng
            big = ((pct >= 2.0) & (pos >= 0.7)).fillna(False).sum()
            out[f'{prefix}大阳次数'] = int(big)
    except Exception:
        out[f'{prefix}大阳次数'] = 0
    return out

def fetch_hs300_data(end_date_str):
    """沪深300多源：东财指数 → 新浪指数 → 510300 ETF → 上证"""
    global BENCHMARK_NAME
    start_date = (datetime.now() - timedelta(days=500)).strftime("%Y%m%d")
    end_str = end_date_str.replace('-', '')

    # 1) 东财 000300
    try:
        df = ak.index_zh_a_hist(symbol="000300", period="daily", start_date=start_date, end_date=end_str)
        df = _normalize_index_df(df)
        if df is not None and len(df) >= 60:
            BENCHMARK_NAME = "沪深300"
            print("  ✅ 基准指数: 沪深300 (000300 东财)")
            return df
    except Exception as e:
        print(f"  ⚠️ HS300 东财失败: {type(e).__name__}: {str(e)[:60]}")

    # 2) 新浪 sh000300
    try:
        df = ak.stock_zh_index_daily(symbol="sh000300")
        df = _normalize_index_df(df)
        if df is not None and len(df) >= 60:
            BENCHMARK_NAME = "沪深300"
            print("  ✅ 基准指数: 沪深300 (sh000300 新浪)")
            return df
    except Exception as e:
        print(f"  ⚠️ HS300 新浪失败: {type(e).__name__}: {str(e)[:60]}")

    # 3) ETF 510300
    try:
        df = ak.stock_zh_a_hist(symbol="510300", period="daily", start_date=start_date, end_date=end_str, adjust="qfq")
        if df is not None and not df.empty:
            df = df.rename(columns={'日期': 'date', '开盘': 'open', '最高': 'high', '最低': 'low', '收盘': 'close', '成交量': 'volume'})
            df = clean_numeric(df)
            if df is not None and len(df) >= 60:
                BENCHMARK_NAME = "沪深300ETF"
                print("  ✅ 基准指数降级: 沪深300ETF (510300)")
                return df
    except Exception as e:
        print(f"  ⚠️ 510300 失败: {type(e).__name__}: {str(e)[:60]}")

    # 4) 上证
    df = fetch_index_hist("000001", end_date_str, "上证指数")
    if df is not None and len(df) >= 60:
        BENCHMARK_NAME = "上证指数"
        print("  ⚠️ 基准指数降级: 上证指数")
        return df

    print("  ❌ 基准指数全部失败，RS 将为 0")
    BENCHMARK_NAME = "未知"
    return None

def fetch_hist_with_cache(code, end_date_str):
    cache_file = CACHE_DIR / f"{code}_{end_date_str.replace('-', '')}_qfq.pkl"
    with CACHE_LOCK:
        if cache_file.exists():
            try:
                age_h = (datetime.now().timestamp() - cache_file.stat().st_mtime) / 3600.0
                if age_h < CACHE_TTL_HOURS:
                    with open(cache_file, 'rb') as f:
                        df = pickle.load(f)
                    df2 = clean_numeric(df)
                    if df2 is not None and not df2.empty and 'date' in df2.columns:
                        try:
                            last_d = pd.to_datetime(df2['date']).max()
                            end_d = pd.to_datetime(end_date_str)
                            if abs((end_d - last_d).days) <= 10:
                                return df2, None
                        except Exception:
                            return df2, None
                try:
                    cache_file.unlink(missing_ok=True)
                except Exception:
                    pass
            except Exception:
                pass
    df = fetch_hist_with_fallback(code, end_date_str)
    if df is not None and len(df) >= MIN_HIST_DAYS:
        df = clean_numeric(df)
        try:
            with CACHE_LOCK:
                with open(cache_file, 'wb') as f:
                    pickle.dump(df, f)
        except Exception:
            pass
        return df, None
    return None, f'{code} 数据获取失败（AkShare+Baostock均失败）'

def batch_fetch_all_hist(codes, end_date_str):
    hist_dict = {}
    errors = []
    print(f"🚀 拉取 {len(codes)} 只股票历史数据（workers={MAX_WORKERS}）...")
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        future_to_code = {executor.submit(fetch_hist_with_cache, code, end_date_str): code for code in codes}
        for future in (tqdm(as_completed(future_to_code), total=len(codes), desc="📥 下载K线") if USE_TQDM else as_completed(future_to_code)):
            code = future_to_code[future]
            try:
                hist, err = future.result()
                if hist is not None:
                    hist_dict[code] = hist
                else:
                    errors.append(f"{code}: {err}")
            except Exception as e:
                errors.append(f"{code}: {str(e)[:100]}")
    print(f"✅ 数据拉取完成！成功 {len(hist_dict)} 只 | 失败 {len(errors)} 只")
    return hist_dict, errors

spot_cache = None
def get_all_spot_data():
    global spot_cache
    if spot_cache is not None and not getattr(spot_cache, 'empty', True):
        return spot_cache
    try:
        df = ak.stock_zh_a_spot_em()
        if df is not None and not df.empty:
            df = df.copy()
            df.columns = [str(c).strip() for c in df.columns]
            spot_cache = df
            return spot_cache
        print("  ⚠️ 实时行情返回空表，不缓存，稍后可重试")
        return pd.DataFrame()
    except Exception as e:
        print(f"  ⚠️ 实时行情获取失败: {type(e).__name__}: {str(e)[:80]}")
        return pd.DataFrame()

def build_spot_lookup(spot_df):
    if spot_df is None or getattr(spot_df, 'empty', True):
        return {}
    df = spot_df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    code_col = next((c for c in df.columns if c in ('代码', 'code')), None)
    if code_col is None:
        return {}
    lookup = {}
    for _, r in df.iterrows():
        key = re.sub(r'\D', '', str(r[code_col])).zfill(6)
        if key:
            lookup[key] = r
    return lookup

def fetch_today_quote(code, spot_df, spot_lookup=None):
    empty = {
        '今日涨跌幅': 0.0, '今日开盘价': 0.0, '今日收盘价': 0.0, '今日成交量': 0,
        '昨收': 0.0, '今开': 0.0, '今日最高': 0.0, '今日最低': 0.0,
        '今日振幅': 0.0, '量比': 0.0, '成交额': 0.0, '换手率': 0.0, '今日均价': 0.0,
        '内盘': 0.0, '外盘': 0.0, '委比': 0.0,
    }
    symbol = get_akshare_symbol(code).zfill(6)
    r = None
    if spot_lookup is not None:
        r = spot_lookup.get(symbol)
    elif spot_df is not None and not getattr(spot_df, 'empty', True):
        df = spot_df
        colmap = {str(c).strip(): c for c in df.columns}
        code_col = colmap.get('代码') or colmap.get('code')
        if code_col is not None:
            s = df[code_col].astype(str).str.replace(r'\D', '', regex=True).str.zfill(6)
            hit = df[s == symbol]
            if not hit.empty:
                r = hit.iloc[0]
    if r is None:
        return empty

    def g(*names, default=0.0):
        for n in names:
            try:
                if isinstance(r, dict):
                    if n not in r:
                        continue
                    val = r[n]
                else:
                    if n not in getattr(r, 'index', []):
                        continue
                    val = r[n]
                if pd.notna(val):
                    return float(val)
            except Exception:
                continue
        return default

    open_p = g('今开', '开盘')
    high_p = g('最高')
    low_p = g('最低')
    close_p = g('最新价', '收盘')
    pre_close = g('昨收')
    pct = g('涨跌幅', 'pct')
    vol = g('成交量', 'volume', '成交手')
    amount = g('成交额', 'amount')
    amp = g('振幅')
    if amp == 0 and pre_close > EPSILON:
        amp = (high_p - low_p) / pre_close * 100
    vol_ratio = g('量比')
    turnover = g('换手率', '换手')

    avg_price = 0.0
    if amount > 0 and vol > 0:
        avg_price = amount / (vol * 100 + EPSILON)
        if avg_price < 0.05 or (close_p > 0 and (avg_price > close_p * 5 or avg_price < close_p * 0.2)):
            avg_price = (high_p + low_p + close_p) / 3 if close_p > 0 else close_p
    elif close_p > 0:
        avg_price = (high_p + low_p + close_p) / 3

    # 东财 spot：成交量/内外盘为「手」，成交额为「元」→ 统一展示 万手 / 亿
    inner = g('内盘', '内盘量')
    outer = g('外盘', '外盘量')
    weibi = g('委比', '委比%')
    return {
        '今日涨跌幅': round(pct, 2),
        '今日开盘价': round(open_p, 2),
        '今日收盘价': round(close_p, 2),
        '今日成交量': round(vol / 1e4, 4) if vol else 0.0,  # 手→万手
        '昨收': round(pre_close, 2),
        '今开': round(open_p, 2),
        '今日最高': round(high_p, 2),
        '今日最低': round(low_p, 2),
        '今日振幅': round(amp, 2),
        '量比': round(vol_ratio, 2),
        '成交额': round(amount / 1e8, 4) if amount else 0.0,  # 元→亿
        '换手率': round(turnover, 2),
        '今日均价': round(avg_price, 2),
        '内盘': round(inner / 1e4, 4) if inner else 0.0,  # 手→万手
        '外盘': round(outer / 1e4, 4) if outer else 0.0,
        '委比': round(weibi, 2),
    }

def calc_streak_and_3d(hist):
    out = {'连涨天': 0, '3日涨%': 0.0, '连续3天振幅%': 0.0, '连续3天均价': 0.0}
    if hist is None or len(hist) < 4:
        return out
    h = hist.sort_values('date').reset_index(drop=True)
    closes = h['close'].astype(float).values
    streak = 0
    for i in range(len(closes) - 1, 0, -1):
        if closes[i] > closes[i - 1]:
            streak += 1
        else:
            break
    out['连涨天'] = int(streak)

    last3 = h.tail(3)
    if len(last3) == 3:
        c0 = float(last3['close'].iloc[0])
        c2 = float(last3['close'].iloc[-1])
        out['3日涨%'] = round((c2 / (c0 + EPSILON) - 1) * 100, 2)
        amps = []
        start_i = len(h) - 3
        for i in range(start_i, len(h)):
            prev_c = float(h['close'].iloc[i - 1])
            hi = float(h['high'].iloc[i])
            lo = float(h['low'].iloc[i])
            if prev_c > EPSILON:
                amps.append((hi - lo) / prev_c * 100)
        out['连续3天振幅%'] = round(float(np.mean(amps)), 2) if amps else 0.0
        out['连续3天均价'] = round(float(last3['close'].mean()), 2)
    return out

# ================== 技术指标算法 ==================

def detect_intraday_limit_up(hist, code, name=""):
    if hist is None or len(hist) < 2:
        return '否', False
    last = hist.iloc[-1]
    prev_close = float(hist['close'].iloc[-2])
    if prev_close <= EPSILON:
        return '否', False
    
    code6 = re.sub(r'\D', '', str(code)).zfill(6)
    name_u = str(name).upper()
    if 'ST' in name_u or '*ST' in name_u:
        limit_pct = 0.05
    elif code6.startswith(('30', '68')):
        limit_pct = 0.20
    elif code6.startswith(('4', '8')):
        limit_pct = 0.30
    else:
        limit_pct = 0.10
    limit_price = prev_close * (1.0 + limit_pct)
    high = float(last['high'])
    close = float(last['close'])
    tol = limit_price * 0.995
    touch = high >= tol
    seal = close >= tol
    if seal:
        return '收盘涨停', True
    if touch:
        return '盘中触及涨停', True
    return '否', False

def backfill_volume_fields_from_hist(today, hist):
    if hist is None or len(hist) < 6:
        return today
    last = hist.iloc[-1]
    vol_raw = float(last['volume']) if 'volume' in hist.columns else 0.0
    amount_raw = float(last['amount']) if 'amount' in hist.columns and pd.notna(last.get('amount', None)) else 0.0
    close_px = float(last['close']) if float(last.get('close', 0) or 0) > 0 else 0.0

    if float(today.get('今日成交量', 0) or 0) <= 0 and vol_raw > 0:
        today['今日成交量'] = round(vol_raw / 1e6, 4)

    if float(today.get('成交额', 0) or 0) <= 0:
        if amount_raw > 0:
            today['成交额'] = round(amount_raw / 1e8, 4)
        elif vol_raw > 0 and close_px > 0:
            today['成交额'] = round(close_px * vol_raw / 1e8, 4)

    if float(today.get('换手率', 0) or 0) <= 0:
        for col in ('turnover_rate', 'turn'):
            if col in hist.columns and float(last.get(col, 0) or 0) > 0:
                today['换手率'] = round(float(last[col]), 2)
                break

    if float(today.get('量比', 0) or 0) <= 0 and vol_raw > 0:
        vol_ma5 = float(hist['volume'].iloc[-6:-1].mean())
        if vol_ma5 > EPSILON:
            today['量比'] = round(vol_raw / vol_ma5, 2)
    return today

def precompute_indicators(hist):
    df = hist.copy()
    df['ma5'] = df['close'].rolling(5).mean()
    df['ma10'] = df['close'].rolling(10).mean()
    df['ma20'] = df['close'].rolling(20).mean()
    return df

def calculate_adx(hist, period=14):
    if len(hist) < period * 2 + 1:
        return 0.0
    high, low, close = hist['high'], hist['low'], hist['close']
    tr = pd.concat([high - low, (high - close.shift()).abs(), (low - close.shift()).abs()], axis=1).max(axis=1)
    plus_dm = high.diff().where((high.diff() > -low.diff()) & (high.diff() > 0), 0)
    minus_dm = (-low.diff()).where((-low.diff() > high.diff()) & (-low.diff() > 0), 0)
    atr = tr.ewm(alpha=1 / period, adjust=False).mean()
    atr_safe = atr.where(atr > EPSILON, EPSILON)
    plus_di = 100 * plus_dm.ewm(alpha=1 / period, adjust=False).mean() / atr_safe
    minus_di = 100 * minus_dm.ewm(alpha=1 / period, adjust=False).mean() / atr_safe
    dx = 100 * abs(plus_di - minus_di) / (plus_di + minus_di + EPSILON)
    adx = dx.ewm(alpha=1 / period, adjust=False).mean()
    return round(adx.iloc[-1], 2)

def detect_technical_patterns(hist):
    if len(hist) < 20:
        return '无明显形态'
    patterns = []
    if hist['ma5'].iloc[-1] > hist['ma10'].iloc[-1] and hist['ma5'].iloc[-2] <= hist['ma10'].iloc[-2]:
        patterns.append("MA5 金叉 MA10")
    if len(hist) >= 21 and hist['close'].iloc[-1] > hist['high'].iloc[-21:-1].max() * 1.02:
        patterns.append("突破 20 日高点")
    vol_ma5 = hist['volume'].rolling(5).mean()
    if hist['volume'].iloc[-1] > vol_ma5.iloc[-1] * 2 and hist['close'].iloc[-1] > hist['close'].iloc[-2] * 1.015:
        patterns.append("放量突破")
    if (hist['close'].tail(3) > hist['close'].tail(3).shift()).sum() >= 2:
        patterns.append("近期连阳")
    if hist['close'].iloc[-1] > hist['ma20'].iloc[-1]:
        patterns.append("站上 MA20")
    return " | ".join(patterns) if patterns else "无明显形态"

def detect_kline_patterns(hist):
    if len(hist) < 3:
        return '数据不足'
    prev2 = hist.iloc[-3]
    prev1 = hist.iloc[-2]
    curr = hist.iloc[-1]
    def body_info(row):
        body = abs(row['close'] - row['open'])
        is_yang = row['close'] > row['open']
        is_yin = row['close'] < row['open']
        return body, is_yang, is_yin
    b2, yang2, yin2 = body_info(prev2)
    b1, yang1, yin1 = body_info(prev1)
    bc, yangc, yinc = body_info(curr)
    avg_body = (hist['close'] - hist['open']).abs().tail(20).mean() if len(hist) >= 20 else (b2 + b1 + bc) / 3
    if avg_body < EPSILON:
        avg_body = EPSILON
    def is_large(body): return body > avg_body * 1.5
    def is_small(body): return body < avg_body * 0.5
    def is_doji(body): return body < avg_body * 0.15
    gap_up = curr['low'] > prev1['high']
    prev1_mid = (prev1['open'] + prev1['close']) / 2
    prev2_mid = (prev2['open'] + prev2['close']) / 2
    strong_signals = []
    normal_signals = []
    lower_shadow_c = min(curr['open'], curr['close']) - curr['low']
    upper_shadow_c = curr['high'] - max(curr['open'], curr['close'])
    if is_small(bc) and lower_shadow_c > bc * 2 and upper_shadow_c < bc * 0.5:
        strong_signals.append("锤头线 (看涨)" if yangc else "上吊线 (警惕)")
    if is_small(bc) and upper_shadow_c > bc * 2 and lower_shadow_c < bc * 0.5:
        normal_signals.append("流星线 (看跌)" if yinc else "倒锤头")
    if is_doji(bc):
        normal_signals.append("十字星")
    if is_large(bc) and yangc and lower_shadow_c < bc * 0.2 and upper_shadow_c < bc * 0.2:
        strong_signals.append("光脚大阳")
    elif is_large(bc) and yinc and lower_shadow_c < bc * 0.2 and upper_shadow_c < bc * 0.2:
        strong_signals.append("光头大阴")
    if yin1 and yangc and bc > b1 * 1.2 and curr['close'] > prev1['open'] and curr['open'] < prev1['close']:
        strong_signals.append("看涨吞没")
    if yang1 and yinc and bc > b1 * 1.2 and curr['close'] < prev1['open'] and curr['open'] > prev1['close']:
        strong_signals.append("看跌吞没")
    if (curr['high'] <= max(prev1['open'], prev1['close']) and curr['low'] >= min(prev1['open'], prev1['close'])):
        if is_doji(bc):
            strong_signals.append("十字孕线")
        elif yang1 and yinc:
            normal_signals.append("看跌孕线")
        elif yin1 and yangc:
            normal_signals.append("看涨孕线")
    if yang1 and yinc and gap_up and curr['close'] < prev1_mid and curr['open'] > prev1['close']:
        strong_signals.append("乌云盖顶")
    if yin1 and yangc and curr['open'] < prev1['low'] and curr['close'] > prev1_mid:
        strong_signals.append("刺透形态")
    if (yin2 and is_large(b2) and is_small(b1) and yangc and is_large(bc) and curr['close'] > prev2_mid):
        strong_signals.append("早晨之星 (见底)")
    if (yang2 and is_large(b2) and is_small(b1) and yinc and is_large(bc) and curr['close'] < prev2_mid):
        strong_signals.append("黄昏之星 (见顶)")
    final_list = []
    seen = set()
    for sig in strong_signals + normal_signals:
        if sig not in seen:
            final_list.append(sig)
            seen.add(sig)
            if len(final_list) >= 3:
                break
    if not final_list:
        if yangc:
            final_list.append("小阳线" if bc <= avg_body * 1.2 else "中阳线")
        elif yinc:
            final_list.append("小阴线" if bc <= avg_body * 1.2 else "中阴线")
        else:
            final_list.append("震荡")
    return " | ".join(final_list)

def check_macd_golden_cross(hist):
    if len(hist) < 35:
        return '无'
    ema12 = hist['close'].ewm(span=12, adjust=False).mean()
    ema26 = hist['close'].ewm(span=26, adjust=False).mean()
    dif = ema12 - ema26
    dea = dif.ewm(span=9, adjust=False).mean()
    cross = (dif.shift(1) < dea.shift(1)) & (dif > dea)
    if cross.tail(5).any():
        return '是（最近金叉）'
    return '否'

def calculate_vwap_cost(hist):
    if len(hist) < 20:
        return 0.0
    recent = hist.tail(20)
    vol_sum = recent['volume'].sum()
    if vol_sum < EPSILON:
        return round(recent['close'].iloc[-1], 2)
    vwap = (recent['close'] * recent['volume']).sum() / vol_sum
    return round(vwap, 2)

def calculate_profit_pct(close, avg_cost):
    if avg_cost <= EPSILON:
        return 0.0
    return round((close - avg_cost) / avg_cost * 100, 2)

def calculate_support_resistance(hist):
    if len(hist) < 10:
        return 0, 0, 0, 0
    return (round(hist['low'].tail(10).min(), 2), round(hist['high'].tail(10).max(), 2),
            round(hist['low'].tail(5).min(), 2), round(hist['high'].tail(5).max(), 2))

def check_valid_breakout(hist, code=""):
    if len(hist) < 12:
        return 0
    market_type = detect_market_type(code)
    min_pct = MARKET_CONFIG[market_type]['breakout_pct']
    df = hist.tail(61).copy()
    df['range'] = df['high'] - df['low']
    df['range'] = df['range'].replace(0, np.nan).fillna(EPSILON)
    df['close_pos'] = (df['close'] - df['low']) / df['range']
    df['vol_ma5'] = df['volume'].rolling(5).mean()
    df['pct'] = (df['close'] / df['close'].shift(1) - 1) * 100
    df = df.iloc[1:]
    valid = (df['pct'] >= min_pct) & (df['close_pos'] >= 0.85) & (df['volume'] > df['vol_ma5'] * 1.5)
    return int(valid.sum())

def calculate_chip_efficiency(hist, code=""):
    if len(hist) < 21:
        return 0.0
    market_type = detect_market_type(code)
    limit_pct = MARKET_CONFIG[market_type]['limit_pct']
    df = hist.tail(21).copy()
    df['is_up'] = df['close'] >= df['open']
    df['is_limit'] = (df['close'] / df['close'].shift(1) - 1) >= limit_pct
    df = df.iloc[1:]
    if df.empty:
        return 0.0
    max_consec = 0
    current = 0
    for is_limit in df['is_limit']:
        if bool(is_limit):
            current += 1
            max_consec = max(max_consec, current)
        else:
            current = 0
    if max_consec >= 3:
        return 15.0
    up_vol = df[df['is_up']]['volume'].sum()
    down_vol = df[~df['is_up']]['volume'].sum()
    if down_vol < EPSILON:
        return 15.0 if up_vol > EPSILON else 0.0
    ratio = up_vol / down_vol
    score = 15 * (1 - np.exp(-ratio / 3))
    return round(min(score, 15.0), 2)

def calculate_obv_trend(hist):
    """OBV 趋势：5日均值 vs 前15日均值。用相对变化率，避免 OBV 为负时乘法阈值反转。"""
    if len(hist) < 20:
        return '震荡'
    close_change = hist['close'].diff().fillna(0)
    obv_change = np.where(close_change > 0, hist['volume'], np.where(close_change < 0, -hist['volume'], 0))
    obv = np.cumsum(obv_change)
    if len(obv) < 20:
        return '震荡'
    obv_5 = float(np.mean(obv[-5:]))
    obv_prev = float(np.mean(obv[-20:-5]))  # 不重叠窗口
    # 差值法：正负 OBV 阈值方向一致（±1% 为震荡带）
    diff_pct = (obv_5 - obv_prev) / (abs(obv_prev) + EPSILON)
    if diff_pct > 0.01:
        return '上升'
    if diff_pct < -0.01:
        return '下降'
    return '震荡'

def check_ma_structure(hist):
    if len(hist) < 30:
        return False
    ma5 = hist['close'].rolling(5).mean()
    ma10 = hist['close'].rolling(10).mean()
    ma20 = hist['close'].rolling(20).mean()
    curr_ma5, curr_ma10, curr_ma20 = ma5.iloc[-1], ma10.iloc[-1], ma20.iloc[-1]
    close = hist['close'].iloc[-1]
    trend_up = curr_ma5 > ma5.iloc[-6] * 1.005 if len(ma5) >= 6 else False
    aligned = (curr_ma5 > curr_ma10 * 1.005) and (curr_ma10 > curr_ma20 * 1.005)
    price_ok = close > curr_ma5 * 0.99
    return aligned and price_ok and trend_up

def get_risk_control(hist, code=""):
    if len(hist) < 20:
        return 0.0
    market_type = detect_market_type(code)
    config = MARKET_CONFIG[market_type]
    h_l = hist['high'] - hist['low']
    h_c = abs(hist['high'] - hist['close'].shift())
    l_c = abs(hist['low'] - hist['close'].shift())
    tr = pd.concat([h_l, h_c, l_c], axis=1).max(axis=1)
    # 与 calculate_adx 统一：Wilder SMMA（ewm alpha=1/14），非 SMA
    atr = float(tr.ewm(alpha=1 / 14, adjust=False).mean().iloc[-1])
    close = hist['close'].iloc[-1]
    if close < EPSILON:
        return 0.0
    daily_vol = atr / close
    atr_mult = config['atr_mult_high'] if daily_vol > 0.05 else config['atr_mult_mid'] if daily_vol > 0.03 else config['atr_mult_low']
    low_10 = float(hist['low'].tail(10).min())
    stop_loss = max(close - atr_mult * atr, low_10)
    stop_loss = max(stop_loss, close * (1 - config['max_drop']))
    if stop_loss >= close:
        stop_loss = close * (1 - config['max_drop'])
    return round(max(stop_loss, 0), 2)

def calculate_relative_strength(hist, hs300_df, window=60):
    if hist is None or hs300_df is None or len(hist) < 20 or len(hs300_df) < 20:
        return 0.0, 0
    h = hist.copy()
    idx = hs300_df.copy()
    h['date'] = pd.to_datetime(h['date'], errors='coerce').dt.tz_localize(None).dt.normalize()
    idx['date'] = pd.to_datetime(idx['date'], errors='coerce').dt.tz_localize(None).dt.normalize()
    h = h.dropna(subset=['date', 'close']).sort_values('date').drop_duplicates('date', keep='last')
    idx = idx.dropna(subset=['date', 'close']).sort_values('date').drop_duplicates('date', keep='last')
    merged = pd.merge(
        h[['date', 'close']].rename(columns={'close': 'close_s'}),
        idx[['date', 'close']].rename(columns={'close': 'close_i'}),
        on='date', how='inner'
    )
    need = max(20, window // 2)
    if len(merged) < need:
        return 0.0, 0
    merged = merged.tail(min(window, len(merged)))
    s0, s1 = float(merged['close_s'].iloc[0]), float(merged['close_s'].iloc[-1])
    i0, i1 = float(merged['close_i'].iloc[0]), float(merged['close_i'].iloc[-1])
    if s0 <= EPSILON or i0 <= EPSILON:
        return 0.0, 0
    stock_ret = s1 / s0
    index_ret = i1 / i0
    if abs(index_ret) < EPSILON:
        return 0.0, 0
    rs = stock_ret / index_ret
    score = 15 if rs > 1.2 else 10 if rs > 1.1 else 5 if rs > 1.0 else 0 if rs > 0.9 else -5
    return round(float(rs), 3), score

def calculate_rsi(hist, period=14):
    if len(hist) < period + 2:
        return 50.0
    delta = hist['close'].diff()
    gain = delta.clip(lower=0).rolling(period).mean()
    loss = (-delta.clip(upper=0)).rolling(period).mean()
    rs = gain / (loss + EPSILON)
    rsi = 100 - (100 / (1 + rs))
    val = rsi.iloc[-1]
    if pd.isna(val):
        return 50.0
    return round(float(val), 2)

def calculate_macd_hist_series(hist):
    if len(hist) < 35:
        return None
    ema12 = hist['close'].ewm(span=12, adjust=False).mean()
    ema26 = hist['close'].ewm(span=26, adjust=False).mean()
    dif = ema12 - ema26
    dea = dif.ewm(span=9, adjust=False).mean()
    return dif - dea

def signal_breakout_20d_volume(hist, vol_mult=1.5):
    if len(hist) < 25:
        return '否'
    high_20 = hist['high'].iloc[-21:-1].max()
    last = hist.iloc[-1]
    prev = hist.iloc[-2]
    vol_ma20 = hist['volume'].iloc[-21:-1].mean()
    is_new_high = (float(last['close']) >= float(high_20) * 0.998) or (float(last['high']) >= float(high_20))
    is_vol = float(last['volume']) >= float(vol_ma20) * vol_mult if vol_ma20 > EPSILON else False
    is_yang = float(last['close']) > float(last['open'])
    is_up = float(last['close']) > float(prev['close'])
    return '是' if (is_new_high and is_vol and is_yang and is_up) else '否'

def calculate_kdj(hist, n=9, m1=3, m2=3):
    if len(hist) < n + 2:
        return 50.0, 50.0, 50.0
    low_n = hist['low'].rolling(n).min()
    high_n = hist['high'].rolling(n).max()
    denom = (high_n - low_n)
    rsv = np.where(denom > EPSILON, (hist['close'] - low_n) / denom * 100, 50.0)
    rsv = pd.Series(rsv, index=hist.index).clip(0, 100)
    k = rsv.ewm(alpha=1 / m1, adjust=False).mean()
    d = k.ewm(alpha=1 / m2, adjust=False).mean()
    j = 3 * k - 2 * d
    return round(float(k.iloc[-1]), 2), round(float(d.iloc[-1]), 2), round(float(j.iloc[-1]), 2)

def signal_oversold_rebound(hist):
    if len(hist) < 40:
        return '否'
    rsi = calculate_rsi(hist, 14)
    macd_hist = calculate_macd_hist_series(hist)
    if macd_hist is None:
        return '否'
    bars = macd_hist.tail(6)
    if bars is None or len(bars) < 3:
        return '否'
    shrink_days = 0
    for i in range(1, len(bars)):
        try:
            if bars.iloc[i] < 0 and bars.iloc[i - 1] < 0:
                if abs(float(bars.iloc[i])) < abs(float(bars.iloc[i - 1])) * 0.98:
                    shrink_days += 1
        except Exception:
            continue
    rsi_ok = 25 <= rsi <= 42
    macd_ok = shrink_days >= 3 and float(bars.iloc[-1]) < 0
    kdj_ok = False
    if len(hist) >= 15:
        low_n = hist['low'].rolling(9).min()
        high_n = hist['high'].rolling(9).max()
        rsv = (hist['close'] - low_n) / (high_n - low_n + EPSILON) * 100
        k = rsv.ewm(alpha=1 / 3, adjust=False).mean()
        d = k.ewm(alpha=1 / 3, adjust=False).mean()
        j = 3 * k - 2 * d
        j1, j0 = float(j.iloc[-1]), float(j.iloc[-2])
        j_min_recent = float(j.tail(5).min())
        kdj_ok = (j_min_recent < 15) and (j1 > j0)
    return '是' if (rsi_ok and macd_ok and kdj_ok) else '否'

def calculate_bias(hist, period=20):
    if len(hist) < period + 1:
        return 0.0
    ma = hist['close'].rolling(period).mean().iloc[-1]
    close = float(hist['close'].iloc[-1])
    if ma is None or pd.isna(ma) or abs(float(ma)) < EPSILON:
        return 0.0
    return round((close / float(ma) - 1.0) * 100, 2)

def detect_market_regime(hs300_df, ma_period=60):
    if hs300_df is None or len(hs300_df) < ma_period + 2:
        return '未知', 0.0, 0.0
    df = hs300_df.copy()
    df['date'] = pd.to_datetime(df['date'], errors='coerce').dt.tz_localize(None).dt.normalize()
    df = df.dropna(subset=['date', 'close']).sort_values('date').drop_duplicates('date', keep='last')
    close = df['close'].astype(float)
    ma60 = close.rolling(ma_period).mean()
    ma20 = close.rolling(20).mean()
    last_c = float(close.iloc[-1])
    last_ma60 = float(ma60.iloc[-1]) if not pd.isna(ma60.iloc[-1]) else 0.0
    last_ma20 = float(ma20.iloc[-1]) if not pd.isna(ma20.iloc[-1]) else 0.0
    if last_ma60 <= EPSILON:
        return '未知', last_c, last_ma60
    above60 = last_c > last_ma60 * 1.002
    below60 = last_c < last_ma60 * 0.998
    above20 = last_ma20 > EPSILON and last_c > last_ma20
    ma20_up = False
    if len(ma20.dropna()) >= 3:
        ma20_up = float(ma20.iloc[-1]) > float(ma20.iloc[-3])
    if above60 and above20 and ma20_up:
        regime = '强势'
    elif above60:
        regime = '偏强'
    elif below60 and not above20:
        regime = '弱势'
    elif below60:
        regime = '偏弱'
    else:
        regime = '中性'
    return regime, last_c, round(last_ma60, 2)

def calculate_max_drawdown(hist, window=60):
    if len(hist) < 10:
        return 0.0
    s = hist['close'].tail(window)
    peak = s.cummax()
    dd = (s - peak) / (peak + EPSILON)
    return round(float(dd.min() * 100), 2)

def calculate_volatility(hist, window=20):
    if len(hist) < window + 1:
        return 0.0
    ret = hist['close'].pct_change().tail(window)
    vol = float(ret.std() * (242 ** 0.5) * 100)
    if pd.isna(vol):
        return 0.0
    return round(vol, 2)

def calculate_liquidity_score(hist):
    if 'turnover_rate' not in hist.columns or len(hist) < 20:
        return 0.0, 0
    avg_turn = hist['turnover_rate'].tail(20).mean()
    score = -15 if avg_turn < 0.8 else 0
    return round(avg_turn, 2), score

def calculate_risk_score(row):
    price = float(row.get('最新价', 0) or 0)
    stop = float(row.get('ATR 止损位', 0) or 0)
    if price <= EPSILON or stop <= EPSILON:
        return 0
    risk_ratio = max((price - stop) / price, 0.0)
    if 0.05 <= risk_ratio <= 0.10:
        return 15
    elif 0.03 <= risk_ratio < 0.05:
        return 12
    elif 0.10 < risk_ratio <= 0.15:
        return 10
    elif risk_ratio > 0.15:
        return 5
    return 0

def calculate_smart_scores(row):
    n = row.get('大阳次数', 0)
    s1 = 15 if n >= 3 else 10 if n >= 2 else 5 if n >= 1 else 0
    c = row.get('筹码效率分', 0.0)
    s2 = 15 if c >= 15 else 10 if c >= 10 else 5 if c >= 5 else 0
    adx = row.get('ADX 趋势强度', 0.0)
    s3 = 15 if adx > 25 else 10 if adx > 20 else 5 if adx > 15 else 0
    s4 = 15 if row.get('均线多头', '否') == '是' else 5
    obv_trend = row.get('OBV 趋势', '震荡')
    s5 = 15 if obv_trend == '上升' else 10 if obv_trend == '震荡' else 5
    s6 = calculate_risk_score(row)
    s7 = row.get('RS 得分', 0)
    liquidity_penalty = row.get('流动性扣分', 0)
    total = s1 + s2 + s3 + s4 + s5 + s6 + s7 + liquidity_penalty
    total = max(0, min(100, total))
    pct = total / 100.0
    if pct >= 0.85:
        r, a = "S 级 (极强)", "重仓出击 (60-70%)"
    elif pct >= 0.75:
        r, a = "A 级 (强势)", "分批建仓 (40-50%)"
    elif pct >= 0.65:
        r, a = "B 级 (观察)", "轻仓试盘 (20-30%)"
    elif pct >= 0.50:
        r, a = "C 级 (弱势)", "观望 (<10%)"
    else:
        r, a = "D 级 (风险)", "排除/止损"
    return pd.Series(
        [s1, s2, s3, s4, s5, s6, s7, liquidity_penalty, total, r, a],
        index=['启动得分', '筹码得分', '趋势得分', '共振得分', '资金得分', '风控得分', 'RS 得分', '流动性扣分', '总分', '评级', '操作建议']
    )

def get_definition_sheet():
    data = {
        '指标名称': ['1. 大阳次数', '2. 筹码效率分', '3. ADX 趋势强度', '4. 均线多头', '5. ATR 止损位', '6. OBV 趋势', '7. 相对强度 RS', '8. 20 日均换手率%', '9. 启动得分', '10. 筹码得分', '11. 趋势得分', '12. 共振得分', '13. 资金得分', '14. 风控得分', '15. RS 得分', '16. 流动性扣分', '17. 总分', '18. 评级', '19. 操作建议'],
        '核心定义': ['60 天内有效大涨（市场自适应）', '涨时放量/跌时缩量（含连板奖励，最高15分）', '趋势强度（Wilder ADX严格SMMA）', 'MA5>MA10>MA20且向上', '动态ATR止损位（最严格保护）', '资金流向趋势（含震荡判断）', '相对沪深300的60日强度（日期对齐）', '流动性门槛（<0.8%扣15分）', '爆发力维度', '筹码锁定维度', '趋势纯度维度', '周期共振维度', '资金流向维度', '风险控制维度', '相对大盘强度（RS>1.2得15分）', '流动性惩罚', '综合评分（0-100分硬保护）', '等级划分', '仓位指导'],
        '计算公式': ['主板≥6%、科创≥8%、北交所≥10% + 收盘位置≥85% + 放量1.5倍', '涨跌量比 + 连续涨停≥3次得15分（上限15）', 'Wilder SMMA（初始简单平均，后续(N-1)/N平滑）', 'MA5>MA10*1.005 & MA10>MA20*1.005 & MA5向上', 'MAX(ATR止损, 10日最低) + 最大回撤下限', '5日OBV vs 15日OBV，±1%为震荡', '个股60日涨幅 / HS300 60日涨幅（实际交易日对齐）', '20日平均换手率', '≥3次=15, ≥2次=10, ≥1次=5', '≥15分=15, ≥10分=10, ≥5分=5', '>25=15, >20=10, >15=5', '是=15, 否=5', '上升=15, 震荡=10, 下降=5', '止损比例5-10%=15, 3-5%=12, 10-15%=10, >15%=5', 'RS>1.2=15, >1.1=10, >1.0=5, >0.9=0, ≤0.9=-5', '<0.8%扣15分', 'Sum(7项得分+扣分)，硬限制0-100', 'S≥85%, A≥75%, B≥65%, C≥50%, D<50%', '按评级执行仓位，D级严格止损'],
        '得分区间': ['0~15分', '0~15分（修复上限）', '0~15分', '5/15分', '自动计算', '5/10/15分', '-5~15分（修复对齐）', '-15~0分', '0~15分', '0~15分', '0~15分', '5/15分', '5~15分', '0~15分', '-5~15分', '-15~0分', '0~100分（硬保护）', 'S/A/B/C/D', '重仓/分批/轻仓/观望/排除'],
        '实盘意义': ['捕捉爆发力，过滤假突破', '判断主力筹码锁定度（上限15分防溢出）', '趋势纯度>25为强趋势股', '均线多头排列是中线持仓基础', '给出最严格止损价，保护本金', '资金是否持续流入', '避免弱市跟风股（日期对齐确保准确）', '剔除地量僵尸股', '爆发力越高越值得重仓', '筹码越集中越容易拉升', '趋势越纯越不容易被砸', '多因子共振胜率更高', '资金是股价的先行指标', '止损距离越小越安全', '相对大盘强势是核心选股逻辑（RS>1.2极强）', '流动性不足的股票风险极高', '综合得分越高越值得加仓（0-100硬保护）', 'S/A级立即行动，D级直接排除', '结合当前市场情绪执行仓位']
    }
    return pd.DataFrame(data)

# ================== 主程序 ==================
if __name__ == "__main__":
    print(f"🚀 启动 A股强势股扫描器 v81_MinCredible Final...")
    END_DATE_STR = get_last_trade_day()
    print(f"📅 扫描基准日期：{END_DATE_STR}")
    print(f"📌 {session_data_note(END_DATE_STR)}")
    hs300 = fetch_hs300_data(END_DATE_STR)
    spot_df = get_all_spot_data()
    spot_lookup = build_spot_lookup(spot_df)
    market_regime, hs300_last, hs300_ma60 = detect_market_regime(hs300, 60)
    print(f"🌡 大盘环境（沪深300 vs MA60）：{market_regime} | 点位≈{hs300_last:.2f} MA60≈{hs300_ma60}")

    print("📡 获取上证/深证指数 → 大盘画像...")
    sh_hist = fetch_index_hist("000001", END_DATE_STR, "上证指数")
    sz_hist = fetch_index_hist("399001", END_DATE_STR, "深证成指")
    market_profile_rows = [
        build_market_profile_row(sh_hist, "上证指数", "000001"),
        build_market_profile_row(sz_hist, "深证成指", "399001"),
    ]
    df_market_profile = pd.DataFrame(market_profile_rows)
    print(f"  上证最新={market_profile_rows[0].get('最新价')} MA5={market_profile_rows[0].get('5日均线')} | "
          f"深证最新={market_profile_rows[1].get('最新价')} MA5={market_profile_rows[1].get('5日均线')}")

    input_path = None
    for p in POSSIBLE_INPUTS:
        if p.exists():
            input_path = p
            break
    if not input_path:
        print("❌ 错误：未找到输入文件！")
        sys.exit(1)
    print(f"📂 已找到输入文件：{input_path.name}")
    try:
        input_df = pd.read_excel(input_path, sheet_name=0)
    except Exception as e:
        print(f"❌ 错误：读取Excel失败 - {e}")
        sys.exit(1)
    code_col = next((c for c in input_df.columns if '代码' in str(c).lower() or 'code' in str(c).lower()), None)
    name_col = next((c for c in input_df.columns if '名称' in str(c).lower() or 'name' in str(c).lower()), None)
    if code_col is None:
        print("❌ 错误：输入文件中未找到“代码”列！")
        sys.exit(1)
    input_df.rename(columns={code_col: '股票代码', name_col or '股票名称': '股票名称'}, inplace=True)
    input_df['股票代码'] = input_df['股票代码'].astype(str).str.strip().str.upper()
    input_df = input_df.drop_duplicates(subset=['股票代码'])
    MY_STOCKS = input_df['股票代码'].tolist()
    name_dict = dict(zip(input_df['股票代码'], input_df['股票名称']))
    print(f"📊 共加载 {len(MY_STOCKS)} 只股票，开始扫描...\n")
    hist_dict, fetch_errors = batch_fetch_all_hist(MY_STOCKS, END_DATE_STR)
    errors = list(fetch_errors)
    if not hist_dict:
        print("❌ 无有效数据")
        sys.exit(0)
    print(f"\n⚡ 开始计算技术指标（{len(hist_dict)} 只）...")
    results = []
    volume_records = []
    iterator = tqdm(hist_dict.items(), desc="计算指标", unit="只") if USE_TQDM else hist_dict.items()
    for code, hist in iterator:
        name = name_dict.get(code, "未知")
        if len(hist) < MIN_HIST_DAYS:
            errors.append(f"{code} ({name}): 数据不足{MIN_HIST_DAYS}天")
            continue
        try:
            df_pre = precompute_indicators(hist)
            today = fetch_today_quote(code, spot_df, spot_lookup)
            streak3 = calc_streak_and_3d(hist)
            avg_cost = calculate_vwap_cost(hist)
            last_close = float(hist['close'].iloc[-1])
            last_open = float(hist['open'].iloc[-1]) if 'open' in hist.columns else last_close
            last_high = float(hist['high'].iloc[-1]) if 'high' in hist.columns else last_close
            last_low = float(hist['low'].iloc[-1]) if 'low' in hist.columns else last_close
            
            spot_px = float(today.get('今日收盘价', 0) or 0)
            use_hist_session = is_weekend_session() or spot_px <= EPSILON
            if use_hist_session:
                current_price = last_close
                if float(today.get('昨收', 0) or 0) <= EPSILON and len(hist) >= 2:
                    today['昨收'] = round(float(hist['close'].iloc[-2]), 2)
                if float(today.get('今开', 0) or 0) <= EPSILON:
                    today['今开'] = round(last_open, 2)
                    today['今日开盘价'] = round(last_open, 2)
                if float(today.get('今日最高', 0) or 0) <= EPSILON:
                    today['今日最高'] = round(last_high, 2)
                if float(today.get('今日最低', 0) or 0) <= EPSILON:
                    today['今日最低'] = round(last_low, 2)
                if float(today.get('今日收盘价', 0) or 0) <= EPSILON:
                    today['今日收盘价'] = round(last_close, 2)
                if float(today.get('今日均价', 0) or 0) <= EPSILON:
                    today['今日均价'] = round((last_high + last_low + last_close) / 3.0, 2)
                if float(today.get('今日振幅', 0) or 0) <= EPSILON and len(hist) >= 2:
                    prev_c = float(hist['close'].iloc[-2])
                    if prev_c > EPSILON:
                        today['今日振幅'] = round((last_high - last_low) / prev_c * 100, 2)
                if abs(float(today.get('今日涨跌幅', 0) or 0)) < EPSILON and len(hist) >= 2:
                    prev_c = float(hist['close'].iloc[-2])
                    if prev_c > EPSILON:
                        today['今日涨跌幅'] = round((last_close / prev_c - 1.0) * 100, 2)
            else:
                current_price = spot_px
            today = backfill_volume_fields_from_hist(today, hist)
            limit_up_tag, _ = detect_intraday_limit_up(hist, code, name)
            profit_pct = calculate_profit_pct(current_price, avg_cost)
            stock_mas = calc_stock_mas(hist)
            adx_val = calculate_adx(hist)
            adx_state = classify_adx_state(adx_val)
            bias_bundle = calc_bias_bundle(hist, current_price)
            short_sup, short_res, ultra_sup, ultra_res = calculate_support_resistance(hist)
            macd_cross = check_macd_golden_cross(hist)
            tech_patterns = detect_technical_patterns(df_pre)
            kline_patterns = detect_kline_patterns(hist)
            rs_value, rs_score = calculate_relative_strength(hist, hs300)
            avg_turnover, liquidity_penalty = calculate_liquidity_score(hist)
            stop_loss = get_risk_control(hist, code)
            stop_distance_pct = round((current_price - stop_loss) / (current_price + EPSILON) * 100, 2) if stop_loss > 0 else 0
            stop_state = classify_stop_distance_state(stop_distance_pct)
            rsi_val = calculate_rsi(hist)
            k_val, d_val, j_val = calculate_kdj(hist)
            sig_break = signal_breakout_20d_volume(hist)
            sig_oversold = signal_oversold_rebound(hist)
            max_dd = calculate_max_drawdown(hist, 60)
            vol_ann = calculate_volatility(hist, 20)
            bias20 = calculate_bias(hist, 20)
            
            signal_tags = []
            if sig_break == '是':
                signal_tags.append('20日突破放量')
            if sig_oversold == '是':
                signal_tags.append('超跌蓄势')
            if str(macd_cross).startswith('是'):
                signal_tags.append('MACD金叉')
            signal_tag = '|'.join(signal_tags) if signal_tags else ''
            
            risk_flags = []
            if max_dd <= -20:
                risk_flags.append('深回撤')
            if vol_ann >= 60:
                risk_flags.append('高波动')
            if stop_distance_pct >= 18:
                risk_flags.append('止损过宽')
            if avg_turnover < 0.8:
                risk_flags.append('低流动')
            if bias20 > BIAS_HIGH_THRESHOLD:
                risk_flags.append('乖离偏高')
            if market_regime == '弱势':
                risk_flags.append('弱市')
            risk_tag = '|'.join(risk_flags) if risk_flags else '正常'
            
            strategy_hits = []
            if sig_break == '是':
                strategy_hits.append('趋势突破')
            if sig_oversold == '是':
                strategy_hits.append('超跌蓄势')
            if today.get('量比', 0) and float(today.get('量比', 0) or 0) >= 1.5 and float(today.get('今日涨跌幅', 0) or 0) > 0:
                strategy_hits.append('放量强势')
            strategy_tag = '|'.join(strategy_hits) if strategy_hits else ''
            
            row = {
                '股票代码': code, '股票名称': name, '最新价': round(current_price, 2),
                '大盘环境': market_regime,
                '昨收': today['昨收'], '今开': today['今开'],
                '今日最高': today['今日最高'], '今日最低': today['今日最低'],
                '今日均价': today['今日均价'],
                '今日涨跌幅%': today['今日涨跌幅'], '今日振幅': today['今日振幅'],
                '量比': today['量比'], '成交额': today['成交额'],
                '盘中涨停': limit_up_tag,
                '换手率%': today['换手率'], '今日成交量': today['今日成交量'],
                '内盘': today['内盘'], '外盘': today['外盘'], '委比': today['委比'],
                '连涨天': streak3['连涨天'], '3日涨%': streak3['3日涨%'],
                '连续3天振幅%': streak3['连续3天振幅%'], '连续3天均价': streak3['连续3天均价'],
                '今日开盘价': today['今日开盘价'], '今日收盘价': today['今日收盘价'],
                '平均成本': avg_cost, '收盘获利%': profit_pct,
                '短线支撑位': short_sup, '短线压力位': short_res,
                '超短线支撑位': ultra_sup, '超短线压力位': ultra_res,
                'MACD 金叉': macd_cross, '技术形态': tech_patterns, 'K 线形态': kline_patterns,
                '大阳次数': check_valid_breakout(hist, code),
                '筹码效率分': calculate_chip_efficiency(hist, code),
                'ADX 趋势强度': adx_val,
                'ADX市场状态': adx_state,
                '均线多头': '是' if check_ma_structure(hist) else '否',
                'ATR 止损位': stop_loss, '止损距离%': stop_distance_pct,
                '止损距离状态': stop_state,
                'OBV 趋势': calculate_obv_trend(hist), '相对强度 RS': rs_value,
                'RS 得分': rs_score, '20 日均换手率%': avg_turnover,
                '流动性扣分': liquidity_penalty, '市场类型': detect_market_type(code),
                '基准指数': BENCHMARK_NAME,
                'RSI(14)': rsi_val,
                'KDJ_K': k_val, 'KDJ_D': d_val, 'KDJ_J': j_val,
                '60日最大回撤%': max_dd,
                '20日年化波动%': vol_ann,
                '信号_20日突破放量': sig_break,
                '信号_超跌蓄势': sig_oversold,
                '信号标签': signal_tag,
                '策略命中': strategy_tag,
                '风险标签': risk_tag,
                'MA5': stock_mas['MA5'], 'MA7': stock_mas['MA7'], 'MA10': stock_mas['MA10'],
                'MA20': stock_mas['MA20'], 'MA60': stock_mas['MA60'], 'MA120': stock_mas['MA120'],
                'BIAS5%': bias_bundle['BIAS5%'], 'BIAS5状态': bias_bundle['BIAS5状态'],
                'BIAS10%': bias_bundle['BIAS10%'], 'BIAS10状态': bias_bundle['BIAS10状态'],
                'BIAS20%': bias_bundle['BIAS20%'], 'BIAS20状态': bias_bundle['BIAS20状态'],
                'BIAS60%': bias_bundle['BIAS60%'], 'BIAS60状态': bias_bundle['BIAS60状态'],
            }
            results.append(row)
            if len(hist) >= 20:
                # 放量倍数：用历史K线「股」口径，避免与展示单位(万手)混算
                vol_ma20_shares = float(hist['volume'].tail(20).mean())
                today_vol_shares = float(hist['volume'].iloc[-1])
                volume_ratio = round(today_vol_shares / (vol_ma20_shares + EPSILON), 2) if vol_ma20_shares > 0 else 0.0
                if volume_ratio >= 1.6 and today['今日涨跌幅'] > -2.0:
                    volume_records.append({
                        '股票代码': code, '股票名称': name, '最新价': row['最新价'],
                        '今日涨跌幅%': today['今日涨跌幅'],
                        '今日成交量(万手)': round(today_vol_shares / 1e6, 4),
                        '20日均量(万手)': round(vol_ma20_shares / 1e6, 4),
                        '放量倍数': volume_ratio,
                        'K 线形态': row['K 线形态'], 'ADX 趋势强度': row['ADX 趋势强度'],
                        'OBV 趋势': row['OBV 趋势']
                    })
        except Exception as e:
            errors.append(f"{code} ({name}): 计算异常 {str(e)}")
            
    print("\n" + "=" * 60)
    print(f"✅ 扫描完成！成功：{len(results)} 只 | 失败：{len(errors)} 只")
    print(f"📌 v81 提示：非交易日实时字段可能为0；请看「行情可信」列与池内分位，勿单独依赖周日结果")
    if not results:
        print("⚠️ 警告：没有成功处理任何股票。")
        sys.exit(0)
    df = pd.DataFrame(results)
    scores = df.apply(calculate_smart_scores, axis=1)
    scores.columns = ['启动得分', '筹码得分', '趋势得分', '共振得分', '资金得分', '风控得分', 'RS 得分', '流动性扣分', '总分', '评级', '操作建议']
    df_final = pd.concat([df, scores], axis=1)
    
    if market_regime == '弱势' and '操作建议' in df_final.columns:
        df_final['操作建议'] = df_final['操作建议'].astype(str) + ' | 弱市慎加仓'
    if '风险标签' in df_final.columns and 'BIAS20%' in df_final.columns:
        mask_bias = df_final['BIAS20%'] > BIAS_HIGH_THRESHOLD
        df_final.loc[mask_bias, '操作建议'] = df_final.loc[mask_bias, '操作建议'].astype(str) + ' | 不追高'
    rating_order = {'S 级 (极强)': 0, 'A 级 (强势)': 1, 'B 级 (观察)': 2, 'C 级 (弱势)': 3, 'D 级 (风险)': 4}
    df_final['评级排序'] = df_final['评级'].map(rating_order)
    df_final = df_final.sort_values(['评级排序', '总分'], ascending=[True, False]).drop('评级排序', axis=1)
    
    rename_units = {
        '成交额': '成交额(亿)',
        '今日成交量': '今日成交量(万手)',
        '内盘': '内盘(万手)',
        '外盘': '外盘(万手)',
    }
    df_final = df_final.rename(columns={k: v for k, v in rename_units.items() if k in df_final.columns})

    if '总分' in df_final.columns and len(df_final) >= 1:
        s = pd.to_numeric(df_final['总分'], errors='coerce').fillna(0)
        df_final['池内总分分位%'] = (s.rank(pct=True, method='average') * 100).round(1) if len(df_final) > 1 else 50.0
    else:
        df_final['池内总分分位%'] = 50.0
    if '相对强度 RS' in df_final.columns and len(df_final) >= 1:
        s = pd.to_numeric(df_final['相对强度 RS'], errors='coerce').fillna(0)
        df_final['池内RS分位%'] = (s.rank(pct=True, method='average') * 100).round(1) if len(df_final) > 1 else 50.0
    else:
        df_final['池内RS分位%'] = 50.0
        
    if is_weekend_session():
        df_final['行情可信'] = f'周末沿用{END_DATE_STR}收盘'
    elif '量比' in df_final.columns:
        df_final['行情可信'] = df_final['量比'].apply(
            lambda x: '是' if float(x or 0) > 0 else '否(spot失败,已回退K线收盘)'
        )
    else:
        df_final['行情可信'] = '未知'
        
    print("📈 生成今日放量Top20预筛选...")
    if volume_records:
        df_volume = pd.DataFrame(volume_records)
        df_volume = df_volume.merge(df_final[['股票代码', '总分', '评级']], on='股票代码', how='left')
        df_volume = df_volume.sort_values('放量倍数', ascending=False).head(20)
        df_volume.insert(0, '放量排名', range(1, len(df_volume) + 1))
        df_volume = df_volume[['放量排名', '股票代码', '股票名称', '最新价', '今日涨跌幅%', '今日成交量(万手)', '20日均量(万手)', '放量倍数', '总分', '评级', 'K 线形态', 'ADX 趋势强度']]
    else:
        df_volume = pd.DataFrame(columns=['放量排名', '股票代码', '股票名称', '最新价', '今日涨跌幅%', '今日成交量(万手)', '20日均量(万手)', '放量倍数', '总分', '评级', 'K 线形态', 'ADX 趋势强度'])
        
    timestamp = datetime.now().strftime('%Y%m%d_%H%M')
    output_path = OUTPUT_DIR / f"自选强势股_v81_MinCredible_{timestamp}.xlsx"
    try:
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            one_page_cols = [c for c in [
                '股票代码', '股票名称', '最新价', '今日涨跌幅%', '总分', '评级',
                '操作建议', '信号标签', '策略命中', '盘中涨停', '风险标签',
                '池内总分分位%', '大盘环境', '行情可信'
            ] if c in df_final.columns]
            df_one = df_final[one_page_cols].copy() if one_page_cols else df_final.iloc[:, :6].copy()
            df_one.to_excel(writer, sheet_name='一页纸决策', index=False)

            # 大盘画像：上证/深证独立一页（不塞进个股结果）
            try:
                df_market_profile.to_excel(writer, sheet_name='大盘画像', index=False)
            except Exception:
                pd.DataFrame([{'说明': '大盘画像生成失败'}]).to_excel(writer, sheet_name='大盘画像', index=False)

            # 个股结果去掉误带的上证/深证列（兼容旧缓存行）
            drop_idx_cols = [c for c in df_final.columns if str(c).startswith('上证') or str(c).startswith('深证')]
            if drop_idx_cols:
                df_final = df_final.drop(columns=drop_idx_cols, errors='ignore')

            df_final.to_excel(writer, sheet_name='股票扫描结果', index=False)
            df_volume.to_excel(writer, sheet_name='今日放量Top20', index=False)

            if '信号_20日突破放量' in df_final.columns:
                sig_mask = (df_final['信号_20日突破放量'] == '是')
                if '信号_超跌蓄势' in df_final.columns:
                    sig_mask = sig_mask | (df_final['信号_超跌蓄势'] == '是')
                df_signal = df_final.loc[sig_mask].copy()
                sig_cols = [c for c in [
                    '股票代码', '股票名称', '最新价', '总分', '评级', '大盘环境', '策略命中',
                    '信号标签', '风险标签', '盘中涨停', '量比', 'RSI(14)', 'KDJ_J', 'BIAS20%',
                    '相对强度 RS', '60日最大回撤%', '20日年化波动%'
                ] if c in df_signal.columns]
                if not df_signal.empty and sig_cols:
                    df_signal[sig_cols].to_excel(writer, sheet_name='信号预警', index=False)
                else:
                    pd.DataFrame(columns=sig_cols or ['说明']).to_excel(writer, sheet_name='信号预警', index=False)

            get_definition_sheet().to_excel(writer, sheet_name='指标完全解读手册', index=False)
            df_final['市场类型'].value_counts().to_frame('数量').to_excel(writer, sheet_name='市场统计')
            df_final['评级'].value_counts().to_frame('数量').to_excel(writer, sheet_name='评级分布')
            if errors:
                pd.DataFrame({'错误详情': errors}).to_excel(writer, sheet_name='错误记录', index=False)

            # ----- openpyxl 美化：无筛选 / 居中 / 雅黑8号 / 列宽适中 -----
            try:
                from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
                from openpyxl.formatting.rule import ColorScaleRule
                from openpyxl.utils import get_column_letter

                header_fill = PatternFill(start_color="1B4F72", end_color="1B4F72", fill_type="solid")
                header_font = Font(name="微软雅黑", size=8, bold=True, color="FFFFFF")
                data_font = Font(name="微软雅黑", size=8, color="2C3E50")
                title_font = Font(name="微软雅黑", size=8, bold=True, color="1B4F72")
                thin = Border(
                    left=Side(style='thin', color='D5D8DC'),
                    right=Side(style='thin', color='D5D8DC'),
                    top=Side(style='thin', color='D5D8DC'),
                    bottom=Side(style='thin', color='D5D8DC')
                )
                alt_fill = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
                white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                rating_fills = {
                    'S 级 (极强)': PatternFill(start_color="F9E79F", end_color="F9E79F", fill_type="solid"),
                    'A 级 (强势)': PatternFill(start_color="F5B7B1", end_color="F5B7B1", fill_type="solid"),
                    'B 级 (观察)': PatternFill(start_color="FCF3CF", end_color="FCF3CF", fill_type="solid"),
                    'C 级 (弱势)': PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid"),
                    'D 级 (风险)': PatternFill(start_color="E5E8E8", end_color="E5E8E8", fill_type="solid"),
                }
                risk_fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
                limit_fill = PatternFill(start_color="F1948A", end_color="F1948A", fill_type="solid")
                ok_fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
                weak_fill = PatternFill(start_color="FDEBD0", end_color="FDEBD0", fill_type="solid")
                warn_fill = PatternFill(start_color="FCF3CF", end_color="FCF3CF", fill_type="solid")
                up_font = Font(name="微软雅黑", size=8, color="C0392B", bold=True)
                down_font = Font(name="微软雅黑", size=8, color="196F3D", bold=True)
                center = Alignment(horizontal="center", vertical="center", wrap_text=True)

                for sheetname in writer.book.sheetnames:
                    ws = writer.book[sheetname]
                    try:
                        ws.sheet_view.showGridLines = True
                    except Exception:
                        pass
                    # 关闭自动筛选（不出现下拉箭头）
                    try:
                        ws.auto_filter.ref = None
                    except Exception:
                        pass

                    ws.row_dimensions[1].height = 20
                    if sheetname in ('股票扫描结果', '一页纸决策', '大盘画像', '信号预警'):
                        ws.freeze_panes = 'C2'
                    else:
                        ws.freeze_panes = 'A2'

                    headers = [cell.value for cell in ws[1]] if ws.max_row >= 1 else []

                    for row_idx in range(1, ws.max_row + 1):
                        for col_idx in range(1, ws.max_column + 1):
                            cell = ws.cell(row=row_idx, column=col_idx)
                            cell.border = thin
                            cell.alignment = center
                            if row_idx == 1:
                                cell.fill = header_fill
                                cell.font = header_font
                            else:
                                cell.fill = alt_fill if (row_idx % 2 == 0) else white_fill
                                cell.font = data_font
                                if isinstance(cell.value, float):
                                    cell.number_format = '0.00'

                    # 列宽：整体略宽，避免 ###；状态/结论列更宽
                    for col in ws.columns:
                        col_letter = get_column_letter(col[0].column)
                        max_len = 10
                        for cell in col[:120]:
                            val_str = str(cell.value if cell.value is not None else '')
                            try:
                                byte_len = len(val_str.encode('gbk', errors='ignore'))
                            except Exception:
                                byte_len = len(val_str)
                            max_len = max(max_len, min(byte_len, 48))
                        h = headers[col[0].column - 1] if col[0].column - 1 < len(headers) else ''
                        if isinstance(h, str) and any(k in h for k in ('状态', '建议', '形态', '结论', '标签', '趋势', '解读')):
                            width = min(max(max_len / 1.6 + 4, 14), 36)
                        elif isinstance(h, str) and any(k in h for k in ('均线', '支撑', '压力', '最新', '最高', '最低')):
                            width = min(max(max_len / 1.6 + 3, 12), 16)
                        else:
                            width = min(max(max_len / 1.7 + 3, 11), 20)
                        ws.column_dimensions[col_letter].width = width

                    if ws.max_row < 2:
                        continue

                    def col_idx_of(name):
                        return headers.index(name) + 1 if name in headers else None

                    def col_letter_of(name):
                        i = col_idx_of(name)
                        return get_column_letter(i) if i else None

                    cidx = col_idx_of('评级')
                    if cidx:
                        for r in range(2, ws.max_row + 1):
                            val = str(ws.cell(row=r, column=cidx).value or '')
                            if val in rating_fills:
                                ws.cell(row=r, column=cidx).fill = rating_fills[val]
                                ws.cell(row=r, column=cidx).font = Font(name="微软雅黑", size=8, bold=True)

                    for col_name, lo, mid, hi in [
                        ('总分', 15, 50, 85),
                        ('池内总分分位%', 0, 50, 100),
                        ('相对强度 RS', 0.7, 1.0, 1.3),
                        ('ADX 趋势强度', 15, 25, 40),
                    ]:
                        letter = col_letter_of(col_name)
                        if letter:
                            ws.conditional_formatting.add(
                                f'{letter}2:{letter}{ws.max_row}',
                                ColorScaleRule(
                                    start_type='num', start_value=lo, start_color='F1948A',
                                    mid_type='num', mid_value=mid, mid_color='F9E79F',
                                    end_type='num', end_value=hi, end_color='82E0AA'
                                )
                            )

                    for col_name in ('今日涨跌幅%', '涨跌幅%', '3日涨%', '收盘获利%'):
                        cidx = col_idx_of(col_name)
                        if not cidx:
                            continue
                        for r in range(2, ws.max_row + 1):
                            cell = ws.cell(row=r, column=cidx)
                            try:
                                v = float(cell.value)
                            except Exception:
                                continue
                            if v > 0:
                                cell.font = up_font
                            elif v < 0:
                                cell.font = down_font

                    cidx = col_idx_of('风险标签')
                    if cidx:
                        for r in range(2, ws.max_row + 1):
                            val = str(ws.cell(row=r, column=cidx).value or '')
                            if val and val != '正常':
                                ws.cell(row=r, column=cidx).fill = risk_fill
                            elif val == '正常':
                                ws.cell(row=r, column=cidx).fill = ok_fill

                    cidx = col_idx_of('盘中涨停')
                    if cidx:
                        for r in range(2, ws.max_row + 1):
                            val = str(ws.cell(row=r, column=cidx).value or '')
                            if val and val != '否':
                                cell = ws.cell(row=r, column=cidx)
                                cell.fill = limit_fill
                                cell.font = Font(name="微软雅黑", size=8, bold=True, color="7B241C")

                    for col_name in ('大盘环境', '相对MA60'):
                        cidx = col_idx_of(col_name)
                        if not cidx:
                            continue
                        for r in range(2, ws.max_row + 1):
                            val = str(ws.cell(row=r, column=cidx).value or '')
                            if any(k in val for k in ('强势', '偏强', '站上')):
                                ws.cell(row=r, column=cidx).fill = ok_fill
                            elif any(k in val for k in ('弱势', '偏弱', '跌破')):
                                ws.cell(row=r, column=cidx).fill = weak_fill

                    overbought = ('极度超买', '超买偏强')
                    oversold = ('极度超卖', '超卖偏弱')
                    for col_name in ('ADX市场状态', '止损距离状态', 'BIAS5状态', 'BIAS10状态', 'BIAS20状态', 'BIAS60状态'):
                        cidx = col_idx_of(col_name)
                        if not cidx:
                            continue
                        for r in range(2, ws.max_row + 1):
                            val = str(ws.cell(row=r, column=cidx).value or '')
                            cell = ws.cell(row=r, column=cidx)
                            if any(k in val for k in overbought) or val in ('超强趋势', '剧烈波动/深回撤中'):
                                cell.fill = risk_fill
                            elif any(k in val for k in oversold) or val in ('无趋势', '极低波动盘整'):
                                cell.fill = weak_fill
                            elif val in ('强趋势', '极强趋势', '健康趋势行情', '中性区间'):
                                cell.fill = ok_fill
                            elif val in ('趋势形成中', '低波动平稳期'):
                                cell.fill = warn_fill

                    for name_col in ('股票名称', '指数名称'):
                        cidx = col_idx_of(name_col)
                        if cidx:
                            for r in range(2, ws.max_row + 1):
                                ws.cell(row=r, column=cidx).font = title_font

                    if sheetname in ('大盘画像', '一页纸决策'):
                        for r in range(2, min(ws.max_row + 1, 50)):
                            ws.row_dimensions[r].height = 18

            except Exception as style_e:
                print(f"  ⚠️ Excel美化部分失败（数据已正常写出）: {style_e}")

        # 固定英文名，供 GitHub raw / 微信 Download（避免中文编码与旧文件误链）
        latest_path = OUTPUT_DIR / "latest.xlsx"
        try:
            shutil.copy2(output_path, latest_path)
            print(f"📌 同步固定下载文件：{latest_path}")
        except Exception as ce:
            print(f"⚠️ 复制 latest.xlsx 失败: {ce}")

        print(f"📊 扫描报告成功输出至：{output_path}")
        print(" 📌 输出包含Sheet：一页纸决策 | 大盘画像 | 股票扫描结果 | 今日放量Top20 | 信号预警")
        cols5 = ['股票名称', '股票代码', '评级', '总分']
        for c in ['池内总分分位%', '相对强度 RS', '信号标签', 'K 线形态']:
            if c in df_final.columns:
                cols5.append(c)
        top_stocks = df_final.head(5)[cols5]
        print("\n🏆 排名前 5 的强势股票预览：")
        print(top_stocks.to_string(index=False))

        if not df_volume.empty:
            print("\n🔥 今日放量 Top5 预筛选：")
            print(df_volume.head(5)[['放量排名', '股票名称', '放量倍数', '今日涨跌幅%', '总分', '评级']].to_string(index=False))
        else:
            print("\n🔥 今日无明显放量股票（放量倍数≥1.6）")

    except Exception as e:
        print(f"❌ 保存文件失败：{e}")
        traceback.print_exc()